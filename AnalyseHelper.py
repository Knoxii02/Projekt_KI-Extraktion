import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
import os
import subprocess
import time
import json
import pyperclip
from Prompt import getPrompt

# pip install pyperclip openpyxl

#--------------------------------------------------------------------------------
# Select one of the AI models from below (ai_models list)
ai_model = "Gemini-3-thinking"

# Select the oscilloscope model to create/load the corresponding Excel file
oszi_model = "MSO54B"
#--------------------------------------------------------------------------------


# Dont change anything below this line

prompt = getPrompt(oszi_model)
print(f"Prompt for {oszi_model} copied")
print("\n----------------")
input("Press Enter after pasting the JSON response...\n----------------")

json_string = pyperclip.paste()
print("Reading JSON data from clipboard...")
json_data = json.loads(json_string)


ai_models = [
    "Gemini-3-thinking",
    "Deepseek",
    "Sonnet-4.5",
    "chatGPT-5.1",
    "chatGPT-4o",
]


row_labels = [
    "Versuch",
    "Modell",
    "Bandbreite",
    "Anzahl Kanäle",
    "Samplerate",
    "Speichertiefe",
    "Triggerarten",
    "Vertikale Auflösung",
    "Anzahl Digitalkanäle",
    "Bildschirmgröße",
    "Bildschirmtyp.pixel",
    "Bildschirmtyp.screen_type",
    "Schnittstellen.relevant_interfaces",
    "Schnittstellen.optionale_interfaces",
    "unterstützende serielle Busse",
    "Signalerfassungsrate",
    "segmentierbarer Speicher",
    "Funktionsgenerator",
    "DVM",
    "Counter",
    "Besonderheiten",
    "Abmessungen (L x B x H) (mm)",
    "Gewicht (kg)",
    "Garantie (Jahre)",
    "Artikelnummer",
    "",
    "",
    "",
    "",
    "JSON"
]

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

#with open('json_data.json', 'r', encoding='utf-8') as f:
#    json_data = json.load(f)

json_obj = json_data[0]
minimized_values = []
for key, value in json_obj.items():
    try:
        val = value['value']
        if isinstance(val, list) and key != "pixel" and "Abmessungen" not in key:
            val = sorted(val, key=lambda x: str(x).lower())
        minimized_values.append(val)
    except:
        for subkey, subvalue in value.items():
            val = subvalue['value']
            if isinstance(val, list) and subkey != "pixel":
                val = sorted(val, key=lambda x: str(x).lower())
            minimized_values.append(val)
            

xlsx_file = f"Datasheets/{oszi_model}.xlsx"

def close_excel_file(filename):
    try:
        subprocess.run(['taskkill', '/F', '/IM', 'EXCEL.EXE'], 
                      capture_output=True, timeout=5)
        time.sleep(1)
        print("Closed excel.")
    except:
        pass

if os.path.exists(xlsx_file):
    close_excel_file(xlsx_file)

neue_datei = not os.path.exists(xlsx_file)

if neue_datei:
    wb = Workbook()
    ws_target = wb.active
    ws_target.title = "Target_Values"
    
    for idx, label in enumerate(row_labels[:25], start=1):
        cell = ws_target.cell(row=idx, column=1, value=label)
        cell.border = thin_border
    
    ws_target.cell(row=1, column=2, value="Target Value")
    ws_target.cell(row=1, column=2).border = thin_border
    
    ws_target.column_dimensions['A'].width = 35
    ws_target.column_dimensions['B'].width = 30
    
    for model_name in ai_models:
        ws = wb.create_sheet(title=model_name)
        
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 30
        
        for idx in range(1, 26):
            cell = ws.cell(row=idx, column=1)
            cell.value = f"=Target_Values!A{idx}"
            cell.border = thin_border
        
        for idx in range(1, 26):
            cell = ws.cell(row=idx, column=2)
            cell.value = f"=Target_Values!B{idx}"
            cell.border = thin_border
        
        green_fill = PatternFill(start_color="9AE383", end_color="9AE383", fill_type="solid")
        
        for row in range(2, 26):
            range_string = f"C{row}:Z{row}"
            
            rule_green = FormulaRule(
                formula=[f'AND(C{row}=$B{row}, NOT(ISBLANK(C{row})), NOT(ISBLANK($B{row})))'], 
                fill=green_fill
            )
            ws.conditional_formatting.add(range_string, rule_green)
            
    
    wb.save(xlsx_file)
    print(f"Neue Datei erstellt mit Sheets: Target_Values, {', '.join(ai_models)}")

wb = openpyxl.load_workbook(xlsx_file)

if ai_model in ai_models:
    ws = wb[ai_model]
else:
    print(f"Warnung: AI-Modell '{ai_model}' nicht in der Liste. Verwende erstes verfügbares Modell-Sheet.")
    ws = wb[ai_models[0]]

versuch_nr = 1
col_idx = 3
while ws.cell(row=1, column=col_idx).value is not None:
    col_idx += 1
    versuch_nr += 1

ws.cell(row=1, column=col_idx, value=versuch_nr)
ws.cell(row=1, column=col_idx).border = thin_border

for idx, value in enumerate(minimized_values, start=2):
    cell = ws.cell(row=idx, column=col_idx)
    
    if isinstance(value, list):
        if len(value) == 1:
            try:
                cell.value = float(value[0]) if '.' in str(value[0]) else int(value[0])
            except (ValueError, TypeError):
                cell.value = ", ".join(str(v) for v in value)
        else:
            formatted_value = ", ".join(str(v) for v in value)
            cell.value = formatted_value
        cell.alignment = Alignment(wrap_text=True, vertical='top')
    else:
        cell.value = value
        cell.alignment = Alignment(vertical='top')
    
    cell.border = thin_border

json_cell = ws.cell(row=30, column=col_idx)
json_cell.value = json.dumps(json_obj, ensure_ascii=False, indent=2)
json_cell.alignment = Alignment(wrap_text=True, vertical='top')
json_cell.border = thin_border

col_letter = openpyxl.utils.get_column_letter(col_idx)
ws.column_dimensions[col_letter].width = 30

wb.save(xlsx_file)
print(f"Saved attempt ``{versuch_nr}`` for model: `{ai_model}` in file: `{xlsx_file}`")

abs_path = os.path.abspath(xlsx_file)
os.startfile(abs_path)
